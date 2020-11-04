import openpyxl

path = r"C:\Users\Majster\Desktop\IEPS\PYTHON\SELENIUM\data.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.active

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for row in range(1,rows+1):
    for column in range(1,cols+1):
        print(sheet.cell(row,column).value, end = "        ")
    print()

